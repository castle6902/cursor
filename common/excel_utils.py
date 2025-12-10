from typing import Optional, List, Dict

import pandas as pd
from openpyxl import load_workbook
from xlrd3 import open_workbook

from logger import log


def excel_read_xlsx(file, sheet_name=None, header=None):
    """读取 xlsx 格式文件。"""
    excel = pd.ExcelFile(load_workbook(file), engine="openpyxl")
    sheet_name = sheet_name or excel.sheet_names[0]
    sheet = excel.book[sheet_name]
    df = excel.parse(sheet_name, header=header)

    for item in sheet.merged_cells:
        top_col, top_row, bottom_col, bottom_row = item.bounds
        base_value = item.start_cell.value
        # 1-based index转为0-based index
        top_row -= 1
        top_col -= 1
        # 由于前面的几行被设为了header，所以这里要对坐标进行调整
        if header is not None:
            top_row -= header + 1
            bottom_row -= header + 1
        df.iloc[top_row:bottom_row, top_col:bottom_col] = base_value
    return df


def excel_read_xlsx_with_multi_sheet(
        file: str,
        sheet_names: Optional[List[str]] = None,
        header: Optional[int] = None
) -> Dict[str, pd.DataFrame]:
    """
    读取 xlsx 格式文件，支持指定多个sheet，自动处理合并单元格（填充合并区域值）。

    参数说明：
    --------
    file: str
        Excel文件路径（绝对路径或相对路径）
    sheet_name: Optional[List[str]]
        需读取的sheet名称列表，默认读取所有sheet
    header: Optional[int]
        表头行索引（0-based），None表示无表头（数据从第1行开始）

    返回值：
    ------
    Dict[str, pd.DataFrame]
        key: sheet名称，value: 对应sheet的DataFrame（已处理合并单元格）
    """
    # 1. 加载工作簿（openpyxl引擎，支持合并单元格识别）
    # 注：直接用openpyxl加载，避免pd.ExcelFile二次封装的冗余
    wb = load_workbook(file, data_only=True)  # data_only=True：读取单元格计算后的值（而非公式）

    # 2. 确定要读取的sheet列表（处理默认值）

    target_sheets = []
    if sheet_names:
        for sheet_name in sheet_names:
            if sheet_name in wb.sheetnames:
                target_sheets.append(sheet_name)
            else:
                log.info(f"找不到 sheet {sheet_name} 在文件 {file}")
    else:
        target_sheets = wb.sheetnames


    # 3. 遍历sheet，逐个读取并处理合并单元格
    sheet_dfs = {}
    for sheet_name in target_sheets:
        ws = wb[sheet_name]

        # 3.1 读取sheet为DataFrame（保留原始行结构，不提前处理表头）
        # 先获取所有数据行（跳过表头的逻辑在后续处理）
        data = []
        for row in ws.iter_rows(values_only=True):  # values_only=True：直接获取单元格值（非Cell对象）
            data.append(list(row))
        # 创建原始DataFrame（行数=Excel行数，列数=Excel最大列数）
        df = pd.DataFrame(data)

        # 3.2 处理合并单元格（核心逻辑）
        # 遍历所有合并单元格区域
        for merged_range in ws.merged_cells.ranges:
            # 获取合并区域的边界（openpyxl返回1-based坐标，需转0-based）
            # CellRange格式：如"A1:C3"，bounds=(min_col, min_row, max_col, max_row)
            min_col, min_row, max_col, max_row = merged_range.bounds
            # 转换为0-based索引（行、列均减1）
            min_row_0 = min_row - 1
            max_row_0 = max_row - 1
            min_col_0 = min_col - 1
            max_col_0 = max_col - 1

            # 获取合并区域的"基准值"（合并单元格的左上角第一个单元格值）
            base_value = ws.cell(row=min_row, column=min_col).value  # 这里用1-based坐标取原始值

            # 关键：根据header调整行索引（表头行不参与数据填充）
            if header is not None:
                # 若合并区域在表头之上（无意义，跳过）
                if max_row_0 <= header:
                    continue
                # 若合并区域包含表头：仅填充表头之下的行（表头行保持原样）
                min_row_0 = max(min_row_0, header + 1)  # 确保起始行不小于表头下一行

            # 填充合并区域（除基准单元格外，其余单元格赋值为基准值）
            # 避免重复填充基准单元格（切片时排除min_row_0:min_row_0+1 或 min_col_0:min_col_0+1）
            # 行范围：min_row_0 到 max_row_0（包含）
            # 列范围：min_col_0 到 max_col_0（包含）
            df.iloc[min_row_0:max_row_0 + 1, min_col_0:max_col_0 + 1] = base_value

        # 3.3 设置表头（若指定header）
        if header is not None:
            # 检查header是否超出数据行数
            if header >= len(df):
                raise IndexError(f"header={header}超出数据行数（共{len(df)}行）")
            # 将指定行设为列名，删除原表头行
            df.columns = df.iloc[header]
            df = df.drop(header).reset_index(drop=True)

        # 3.4 存储当前sheet的DataFrame
        sheet_dfs[sheet_name] = df

    # 4. 关闭工作簿（释放资源）
    wb.close()

    return sheet_dfs


def excel_read_xls(file, sheet_name=None, header=None):
    """读取 xls 格式文件。"""
    excel = pd.ExcelFile(open_workbook(file, formatting_info=True), engine="xlrd")
    sheet_name = sheet_name or excel.sheet_names[0]
    sheet = excel.book[sheet_name]
    df = excel.parse(sheet_name, header=header)

    # 0-based index
    for top_row, bottom_row, top_col, bottom_col in sheet.merged_cells:
        base_value = sheet.cell_value(top_row, top_col)
        # 由于前面的几行被设为了header，所以这里要对坐标进行调整
        if header is not None:
            top_row -= header + 1
            bottom_row -= header + 1
        df.iloc[top_row:bottom_row, top_col:bottom_col] = base_value
    return df
