from typing import Dict, Tuple, Any, List

import pandas as pd
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from bussiness.lap_file_parse.config_of_djz import DJZConfig
from bussiness.lap_file_parse.config_of_lcmy import LCMYConfig
from bussiness.lap_file_parse.config_of_ncg import NCGConfig
from bussiness.lap_file_parse.config_of_xcg import XCGConfig
from bussiness.lap_file_parse.config_of_xyng import XYNGConfig
from bussiness.lap_file_parse.lap_dipp_upload_utils import upload_template_file_to_dipp, upload_data_to_dipp
from bussiness.lap_file_parse.model import LapRequestModel, LapFileType, FileInfo
from bussiness.lap_file_parse.service.djz_parser import djz_parser
from bussiness.lap_file_parse.service.lcmy_parser import lcmy_parser
from bussiness.lap_file_parse.service.ncg_parser import ncg_parser
from bussiness.lap_file_parse.service.template_parse import template_file_of_xlsx_parse
from bussiness.lap_file_parse.service.xcg_parser import xcg_parser
from bussiness.lap_file_parse.service.xyng_parser import xyng_parser
from bussiness.lap_file_parse.thread_pools import parse_pool

import os
from concurrent.futures import wait

from bussiness.dipp.dipp_file_utils import download_file
from common.utils.df_chain_utils import PandasChain
from logger import log


# formate_sheet_view 可以用的话，这个就废弃了
def add_border(ws: Worksheet, thin_border: Border) -> None:
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border  # 应用边框


# formate_sheet_view 可以用的话，这个就废弃了
def highlight_error_rows(ws: Worksheet, yellow_fill: PatternFill, header_list: List[str]) -> None:
    error_col_idx = len(header_list)
    if error_col_idx is None:
        log.info("请检查表头是否正确")
        return
    #
    # # 遍历数据行（从第 4 行开始，跳过表头）
    for row in ws.iter_rows(min_row=4):
        # 获取当前行的行号（如第5行、第6行...）
        current_row_num = row[0].row

        # 获取当前行的ERROR列单元格
        # 直接获取第4列（ERROR列）的单元格
        error_cell = ws.cell(row=current_row_num, column=error_col_idx)
        error_value_info = error_cell.value
        # 如果ERROR列非空，对A-D列（1-4列）染色
        if error_value_info not in (None, ""):

            # col_idx 列号， 1 对应 A 列，  2 对应 B 列
            # 整行染色，
            for col_idx in range(1, error_col_idx + 1):  # A(1)到D(4)
                cell = ws.cell(row=current_row_num, column=col_idx)
                cell.fill = yellow_fill


def formate_sheet_view(ws: Worksheet, yellow_fill: PatternFill, header_list: List[str], border: Border,
                       file_type: LapFileType) -> None:
    error_col_idx = len(header_list)
    if error_col_idx is None:
        log.info("请检查表头是否正确")
        return
    # 从第 min_row 行开始才是数据的开始列
    min_row = 4
    if file_type == LapFileType.SHDJZ or file_type == LapFileType.XCG:
        min_row = 5

    # # 遍历数据行（从第 4 行开始，跳过表头）
    for row in ws.iter_rows(min_row=min_row):
        # 获取当前行的行号（如第5行、第6行...）
        current_row_num = row[0].row

        # 获取当前行的ERROR列单元格
        # 直接获取第4列（ERROR列）的单元格
        error_cell = ws.cell(row=current_row_num, column=error_col_idx)
        error_value_info = error_cell.value

        # 如果ERROR列非空，说明需要对 error 列进行染色
        # 对A-D列（1-4列）染色
        if error_value_info not in (None, ""):
            # col_idx 列号， 1 对应 A 列，  2 对应 B 列
            # 整行染色，
            for col_idx in range(1, error_col_idx + 1):  # A(1)到D(4)
                cell = ws.cell(row=current_row_num, column=col_idx)
                cell.fill = yellow_fill
                cell.border = border  # 应用边框
        else:
            for col_idx in range(1, error_col_idx):  # A(1)到D(4)
                cell = ws.cell(row=current_row_num, column=col_idx)
                cell.border = border  # 应用边框


def write_to_template_file(template_file_path: str, file_parse_data: list[tuple[LapFileType, PandasChain]]):
    log.info(f"开始进行 excel　写入")
    # 定义黄色填充样式（RGB颜色：FFFF00）
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 边框线
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    with pd.ExcelWriter(template_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        wb = writer.book

        for file_type, file_df_chain in file_parse_data:
            sheet_name = file_type.value
            # target_sheet = wb[sheet_name]
            log.info(f"开始进行 sheet　写入 {sheet_name}")
            cur_sheet = writer.book[sheet_name]
            startrow = cur_sheet.max_row

            cuf_pd = file_df_chain.df.apply(pd.to_numeric, errors="ignore")
            cuf_pd.to_excel(writer, sheet_name=sheet_name, startrow=startrow, header=False, index=False)

            if file_type == LapFileType.SHDJZ or file_type == LapFileType.XCG:
                # 冻结窗口
                cur_sheet.freeze_panes = "C5"
            else:
                cur_sheet.freeze_panes = "C4"

            # 添加 border
            # add_border(ws=cur_sheet, thin_border=thin_border)

            # 进行行染色
            # if file_type == LapFileType.SHDJZ or file_type == LapFileType.LCMY:
            #     highlight_error_rows(ws=cur_sheet, yellow_fill=yellow_fill,
            #                          header_list=file_df_chain.return_header_list())
            formate_sheet_view(ws=cur_sheet, yellow_fill=yellow_fill, header_list=file_df_chain.return_header_list(),
                               border=thin_border, file_type=file_type)


def multi_thread_parse(lap_type: LapFileType, lap_file_info_list: List[FileInfo], header_of_template: list[str]):
    if lap_type == LapFileType.SHDJZ:
        return djz_parser(lap_file_info_list, header_of_template)
    elif lap_type == LapFileType.XCG:
        return xcg_parser(lap_file_info_list, header_of_template)
    elif lap_type == LapFileType.LCMY:
        return lcmy_parser(lap_file_info_list, header_of_template)
    elif lap_type == LapFileType.XYNG:
        return xyng_parser(lap_file_info_list, header_of_template)
    elif lap_type == LapFileType.NCG:
        return ncg_parser(lap_file_info_list, header_of_template)
    else:
        return ()


# 统一解析url的方法
def parse_file(request_body: LapRequestModel):
    # ==============================================
    # 对文件进行下载，然后解析， 下载下来的格式是
    # [fileInfo, fileInfo]
    # ==============================================

    all_files_info = request_body.parse_all_file_url()
    log.info(f"开始下载文件:  {all_files_info}")
    futures = [
        parse_pool.submit(
            download_file,
            download_file_info.download_url,
            request_body.token,
            download_file_info.save_path
        )
        for i, download_file_info in enumerate(all_files_info)
    ]
    # 阻塞直到所有任务完成（可设置超时）
    wait(futures, return_when="ALL_COMPLETED")
    log.info(f"文件进行解析")
    log.info("开始对需要解析的文件进行分门别类，依次进行解析，优先解析模板文件")

    # ==============================================
    # 对解析的文件进行分类
    # {file_type: 【fileInfo，fileInfo]}
    # ==============================================
    parse_file_info_map = {}
    for e in all_files_info:
        type = e.type
        file_of_one_type = parse_file_info_map.get(type, [])
        file_of_one_type.append(e)
        parse_file_info_map[type] = file_of_one_type

    log.info("开始解析模板文件")
    # ==============================================
    # 根据 fileType 获取所需要的 templateHeader
    # {file_type: [header， header]}
    # ==============================================
    template_file_info = parse_file_info_map.get(LapFileType.Template)[0]
    del parse_file_info_map[LapFileType.Template]
    header_infor = template_file_of_xlsx_parse(file_info=template_file_info,
                                               sheet_name_list=list(parse_file_info_map.keys()))

    log.info(f"解析到需求的 header: {header_infor}")
    log.info("根据不同的类型，开始异步解析加载各种文件")
    # [type, [fileInfo, fileInfo, fileInfo]

    # for file_path in download_file_save_path:
    #     pass

    # multi_thread_parse(type, [LapFileInfo])
    # ========================================================
    # 方式一： 多线程解析
    # ========================================================
    futures = [
        parse_pool.submit(
            multi_thread_parse,
            key,
            value,
            header_infor.get(key)
        )
        for key, value in parse_file_info_map.items()
    ]
    #
    # # 获取所有结果（按完成顺序）

    # wait(futures, return_when="ALL_COMPLETED")
    results = [future.result() for future in futures]
    # ========================================================
    # 写入 excel
    # ========================================================
    log.info(f"解析完成，进行数据合并, 写入 excel")
    write_to_template_file(template_file_path=template_file_info.save_path, file_parse_data=results)
    # ========================================================
    # 推送template.xlsx 到 dipp 当中
    # ========================================================
    log.info(f"写入 excel 完成后，开始进行推送 template.xlsx 模板文件")
    # upload_template_file_to_dipp(url: str, template_save_path, token):
    upload_template_file_to_dipp(url=request_body.upload_template_file_to_dipp(),
                                 template_save_path=template_file_info.save_path,
                                 token=request_body.token)

    # ========================================================
    # 推送 data 到 dipp 当中
    # ========================================================
    log.info(f"推送完成 template.xlsx 模板文件，开始推送数据")
    # def upload_data_to_dipp(url: str, df_chain: PandasChain, token: str):

    task = []
    for file_type, file_df_chain in results:
        #    def column_add_column(self, header: str, default_value: str = ""):
        base_df_chain: PandasChain = (
            file_df_chain.column_add_column(header="xiangmubianma", default_value=request_body.project_code)
            .column_add_column(header="xiangmubianhao", default_value=request_body.project_num)
            .column_add_column(header="baogaobianma", default_value=request_body.report_id)
            .column_add_column(header="shijiandian", default_value=request_body.time_point)
        )
        if file_type == LapFileType.SHDJZ:
            payload = (
                base_df_chain.header_rename_if_exist(DJZConfig.header_map_template_to_dipp)
                .return_dict_without()
            )
            # upload_data_to_dipp(url=request_body.upload_data_url_to_SHDJZ(), payload=payload,
            #                     token=request_body.token)
            future1 = parse_pool.submit(upload_data_to_dipp, request_body.upload_data_url_to_SHDJZ(), payload,
                                        request_body.token)
            task.append(future1)
        elif file_type == LapFileType.XCG:
            payload = (
                base_df_chain.header_rename_if_exist(XCGConfig.header_map_template_to_dipp)
                .return_dict_without()
            )
            # upload_data_to_dipp(url=request_body.upload_data_url_to_XCG(), payload=payload,
            #                     token=request_body.token)
            future2 = parse_pool.submit(upload_data_to_dipp, request_body.upload_data_url_to_XCG(), payload,
                                        request_body.token)
            task.append(future2)
        elif file_type == LapFileType.NCG:
            payload = (
                base_df_chain.header_rename_if_exist(NCGConfig.header_map_template_to_dipp)
                .return_dict_without()
            )
            # upload_data_to_dipp(url=request_body.upload_data_url_to_NCG(), payload=payload,
            #                     token=request_body.token)
            future3 = parse_pool.submit(upload_data_to_dipp, request_body.upload_data_url_to_NCG(), payload,
                                        request_body.token)
            task.append(future3)
        elif file_type == LapFileType.LCMY:
            payload = (
                base_df_chain.header_rename_if_exist(LCMYConfig.header_map_template_to_dipp)
                .return_dict_without()
            )
            # upload_data_to_dipp(url=request_body.upload_data_url_to_LCMY(), payload=payload,
            #                     token=request_body.token)
            future4 = parse_pool.submit(upload_data_to_dipp, request_body.upload_data_url_to_LCMY(), payload,
                                        request_body.token)
            task.append(future4)
        elif file_type == LapFileType.XYNG:

            payload = (
                base_df_chain.header_rename_if_exist(XYNGConfig.header_map_template_to_dipp)
                .return_dict_without()
            )
            # upload_data_to_dipp(url=request_body.upload_data_url_to_XYNG(), payload=payload,
            #                     token=request_body.token)
            future5 = parse_pool.submit(upload_data_to_dipp, request_body.upload_data_url_to_XYNG(), payload,
                                        request_body.token)
            task.append(future5)

        log.info(f"数据推送完成，{file_type}")
    wait(task, return_when="ALL_COMPLETED")  # 阻塞直到所有任务结束
    log.info(f"推送数据完成，全部结束")
